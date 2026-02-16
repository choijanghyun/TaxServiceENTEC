#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
세액공제 환급 계산 시스템 - 테스트 케이스 생성 스크립트
법인(CORP) 10건 + 개인(INC) 10건 = 총 20건
각 케이스별 Excel(.xlsx) + JSON(.json) 페어 생성
"""

import json
import os
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ============================================================
# 법인 테스트 케이스 10건
# ============================================================
CORP_CASES = [
    {
        "id": "TC-CORP-01",
        "title": "중소기업_고용증대_세액공제",
        "desc": "수도권 외 중소기업이 청년 고용을 늘려 고용증대 세액공제를 신청하는 기본 시나리오",
        "basic": {
            "applicant_type": "C", "applicant_name": "주식회사 한빛테크",
            "biz_reg_no": "123-45-67890", "tax_type": "CORP", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "62010",
            "hq_location": "대전광역시 유성구 대학로 99", "capital_zone": "NON_CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 5000000000, "taxable_income": 800000000,
            "computed_tax": 152000000, "paid_tax": 152000000,
            "founding_date": "2018-03-15", "venture_yn": False, "rd_dept_yn": True,
            "claim_reason": "고용증대 세액공제 경정청구", "sincerity_target": False,
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 45.0, "youth_count": 12, "disabled_count": 1,
             "aged_count": 2, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 30, "excluded_count": 3, "total_salary": 2700000000, "social_insurance_paid": 270000000},
            {"year_type": "PREVIOUS", "total_regular": 38.0, "youth_count": 8, "disabled_count": 1,
             "aged_count": 2, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 27, "excluded_count": 2, "total_salary": 2200000000, "social_insurance_paid": 220000000},
        ],
        "deductions": [
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2024", "item_seq": 1,
             "base_amount": 0, "zone_type": "NON_CAPITAL", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": None, "non_taxable_income": 50000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 70000000,
            "withholding_tax": 5000000, "determined_tax": 152000000,
            "dividend_income_total": 0, "foreign_tax_total": 0, "foreign_income_total": 0,
        },
    },
    {
        "id": "TC-CORP-02",
        "title": "중소기업_RD_세액공제_신성장원천",
        "desc": "중소기업이 신성장·원천기술 R&D 세액공제를 신청하는 시나리오",
        "basic": {
            "applicant_type": "C", "applicant_name": "주식회사 바이오젠",
            "biz_reg_no": "234-56-78901", "tax_type": "CORP", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "21100",
            "hq_location": "경기도 성남시 분당구 판교로 123", "capital_zone": "CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 12000000000, "taxable_income": 2000000000,
            "computed_tax": 400000000, "paid_tax": 400000000,
            "founding_date": "2015-07-01", "venture_yn": True, "rd_dept_yn": True,
            "claim_reason": "R&D 세액공제 경정청구 (신성장원천기술)", "sincerity_target": False,
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 80.0, "youth_count": 25, "disabled_count": 0,
             "aged_count": 3, "career_break_count": 1, "north_defector_count": 0,
             "general_count": 51, "excluded_count": 5, "total_salary": 6000000000, "social_insurance_paid": 600000000},
            {"year_type": "PREVIOUS", "total_regular": 75.0, "youth_count": 22, "disabled_count": 0,
             "aged_count": 3, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 50, "excluded_count": 4, "total_salary": 5500000000, "social_insurance_paid": 550000000},
        ],
        "deductions": [
            {"item_category": "RD", "provision": "조특법10", "tax_year": "2024", "item_seq": 1,
             "base_amount": 3000000000, "zone_type": None, "asset_type": None,
             "rd_type": "NEW_GROWTH", "method": "INCREMENT",
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": None, "non_taxable_income": 100000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 180000000,
            "withholding_tax": 10000000, "determined_tax": 400000000,
            "dividend_income_total": 0, "foreign_tax_total": 0, "foreign_income_total": 0,
        },
    },
    {
        "id": "TC-CORP-03",
        "title": "중소기업_투자세액공제",
        "desc": "중소기업이 생산설비 투자에 대한 세액공제를 신청하는 시나리오",
        "basic": {
            "applicant_type": "C", "applicant_name": "주식회사 대한정밀",
            "biz_reg_no": "345-67-89012", "tax_type": "CORP", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "29199",
            "hq_location": "충청남도 천안시 서북구 공단1로 50", "capital_zone": "NON_CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 8000000000, "taxable_income": 1200000000,
            "computed_tax": 228000000, "paid_tax": 228000000,
            "founding_date": "2010-05-20", "venture_yn": False, "rd_dept_yn": False,
            "claim_reason": "생산설비 투자세액공제 경정청구", "sincerity_target": False,
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 60.0, "youth_count": 15, "disabled_count": 2,
             "aged_count": 5, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 38, "excluded_count": 4, "total_salary": 3600000000, "social_insurance_paid": 360000000},
            {"year_type": "PREVIOUS", "total_regular": 58.0, "youth_count": 14, "disabled_count": 2,
             "aged_count": 5, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 37, "excluded_count": 3, "total_salary": 3400000000, "social_insurance_paid": 340000000},
        ],
        "deductions": [
            {"item_category": "INVESTMENT", "provision": "조특법24", "tax_year": "2024", "item_seq": 1,
             "base_amount": 2000000000, "zone_type": "NON_CAPITAL", "asset_type": "FACILITY",
             "rd_type": None, "method": None,
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": None, "non_taxable_income": 30000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 100000000,
            "withholding_tax": 8000000, "determined_tax": 228000000,
            "dividend_income_total": 0, "foreign_tax_total": 0, "foreign_income_total": 0,
        },
    },
    {
        "id": "TC-CORP-04",
        "title": "중견기업_복합세액공제",
        "desc": "중견기업이 고용증대 + R&D + 투자 세액공제를 복합 신청하는 시나리오",
        "basic": {
            "applicant_type": "C", "applicant_name": "주식회사 세종일렉트로닉스",
            "biz_reg_no": "456-78-90123", "tax_type": "CORP", "tax_year": "2024",
            "corp_size": "MEDIUM", "industry_code": "26410",
            "hq_location": "경기도 화성시 동탄산단1길 30", "capital_zone": "CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 50000000000, "taxable_income": 5000000000,
            "computed_tax": 1100000000, "paid_tax": 1100000000,
            "founding_date": "2005-01-10", "venture_yn": False, "rd_dept_yn": True,
            "claim_reason": "고용증대·R&D·투자 복합 세액공제 경정청구", "sincerity_target": False,
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 350.0, "youth_count": 80, "disabled_count": 5,
             "aged_count": 20, "career_break_count": 3, "north_defector_count": 0,
             "general_count": 242, "excluded_count": 15, "total_salary": 28000000000, "social_insurance_paid": 2800000000},
            {"year_type": "PREVIOUS", "total_regular": 320.0, "youth_count": 65, "disabled_count": 5,
             "aged_count": 18, "career_break_count": 2, "north_defector_count": 0,
             "general_count": 230, "excluded_count": 12, "total_salary": 25000000000, "social_insurance_paid": 2500000000},
        ],
        "deductions": [
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2024", "item_seq": 1,
             "base_amount": 0, "zone_type": "CAPITAL", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
            {"item_category": "RD", "provision": "조특법10", "tax_year": "2024", "item_seq": 1,
             "base_amount": 8000000000, "zone_type": None, "asset_type": None,
             "rd_type": "GENERAL", "method": "CURRENT",
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
            {"item_category": "INVESTMENT", "provision": "조특법24", "tax_year": "2024", "item_seq": 1,
             "base_amount": 5000000000, "zone_type": "CAPITAL", "asset_type": "FACILITY",
             "rd_type": None, "method": None,
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": None, "non_taxable_income": 200000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 500000000,
            "withholding_tax": 30000000, "determined_tax": 1100000000,
            "dividend_income_total": 0, "foreign_tax_total": 0, "foreign_income_total": 0,
        },
    },
    {
        "id": "TC-CORP-05",
        "title": "대기업_RD_세액공제_최저한세",
        "desc": "대기업이 R&D 세액공제 신청 시 최저한세 적용으로 공제 제한이 걸리는 시나리오",
        "basic": {
            "applicant_type": "C", "applicant_name": "주식회사 글로벌반도체",
            "biz_reg_no": "567-89-01234", "tax_type": "CORP", "tax_year": "2024",
            "corp_size": "LARGE", "industry_code": "26110",
            "hq_location": "서울특별시 강남구 테헤란로 152", "capital_zone": "CAPITAL_CONGESTION",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 500000000000, "taxable_income": 80000000000,
            "computed_tax": 17400000000, "paid_tax": 17400000000,
            "founding_date": "1990-03-01", "venture_yn": False, "rd_dept_yn": True,
            "claim_reason": "R&D 세액공제 경정청구 (일반연구·인력개발)", "sincerity_target": False,
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 2500.0, "youth_count": 500, "disabled_count": 30,
             "aged_count": 100, "career_break_count": 20, "north_defector_count": 0,
             "general_count": 1850, "excluded_count": 80, "total_salary": 250000000000, "social_insurance_paid": 25000000000},
            {"year_type": "PREVIOUS", "total_regular": 2450.0, "youth_count": 480, "disabled_count": 28,
             "aged_count": 98, "career_break_count": 18, "north_defector_count": 0,
             "general_count": 1826, "excluded_count": 75, "total_salary": 240000000000, "social_insurance_paid": 24000000000},
        ],
        "deductions": [
            {"item_category": "RD", "provision": "조특법10", "tax_year": "2024", "item_seq": 1,
             "base_amount": 30000000000, "zone_type": None, "asset_type": None,
             "rd_type": "GENERAL", "method": "CURRENT",
             "existing_applied": True, "existing_amount": 500000000, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": None, "non_taxable_income": 2000000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 8000000000,
            "withholding_tax": 500000000, "determined_tax": 17400000000,
            "dividend_income_total": 3000000000, "foreign_tax_total": 200000000, "foreign_income_total": 5000000000,
        },
    },
    {
        "id": "TC-CORP-06",
        "title": "창업중소기업_감면",
        "desc": "수도권 외 창업중소기업이 5년간 세액감면을 적용하는 시나리오",
        "basic": {
            "applicant_type": "C", "applicant_name": "주식회사 스마트팩토리",
            "biz_reg_no": "678-90-12345", "tax_type": "CORP", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "62021",
            "hq_location": "전라북도 전주시 덕진구 혁신로 45", "capital_zone": "NON_CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 2000000000, "taxable_income": 300000000,
            "computed_tax": 39000000, "paid_tax": 39000000,
            "founding_date": "2022-06-01", "venture_yn": True, "rd_dept_yn": True,
            "claim_reason": "창업중소기업 세액감면 경정청구", "sincerity_target": False,
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 20.0, "youth_count": 10, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 10, "excluded_count": 1, "total_salary": 1200000000, "social_insurance_paid": 120000000},
            {"year_type": "PREVIOUS", "total_regular": 12.0, "youth_count": 6, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 6, "excluded_count": 0, "total_salary": 700000000, "social_insurance_paid": 70000000},
        ],
        "deductions": [
            {"item_category": "STARTUP", "provision": "조특법6", "tax_year": "2024", "item_seq": 1,
             "base_amount": 300000000, "zone_type": "NON_CAPITAL", "asset_type": None,
             "rd_type": None, "method": None,
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": None, "non_taxable_income": 10000000,
            "loss_carryforward_total": 50000000, "interim_prepaid_tax": 15000000,
            "withholding_tax": 2000000, "determined_tax": 39000000,
            "dividend_income_total": 0, "foreign_tax_total": 0, "foreign_income_total": 0,
        },
    },
    {
        "id": "TC-CORP-07",
        "title": "수도권과밀억제권역_투자제한",
        "desc": "수도권 과밀억제권역 소재 기업의 투자세액공제 제한 적용 시나리오",
        "basic": {
            "applicant_type": "C", "applicant_name": "주식회사 서울커머스",
            "biz_reg_no": "789-01-23456", "tax_type": "CORP", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "47911",
            "hq_location": "서울특별시 구로구 디지털로26길 111", "capital_zone": "CAPITAL_CONGESTION",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 15000000000, "taxable_income": 1500000000,
            "computed_tax": 285000000, "paid_tax": 285000000,
            "founding_date": "2012-11-01", "venture_yn": False, "rd_dept_yn": False,
            "claim_reason": "투자세액공제 경정청구 (수도권 과밀억제권역)", "sincerity_target": False,
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 100.0, "youth_count": 30, "disabled_count": 2,
             "aged_count": 5, "career_break_count": 1, "north_defector_count": 0,
             "general_count": 62, "excluded_count": 8, "total_salary": 7000000000, "social_insurance_paid": 700000000},
            {"year_type": "PREVIOUS", "total_regular": 95.0, "youth_count": 28, "disabled_count": 2,
             "aged_count": 5, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 60, "excluded_count": 7, "total_salary": 6500000000, "social_insurance_paid": 650000000},
        ],
        "deductions": [
            {"item_category": "INVESTMENT", "provision": "조특법24", "tax_year": "2024", "item_seq": 1,
             "base_amount": 3000000000, "zone_type": "CAPITAL_CONGESTION", "asset_type": "FACILITY",
             "rd_type": None, "method": None,
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": None, "non_taxable_income": 80000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 130000000,
            "withholding_tax": 15000000, "determined_tax": 285000000,
            "dividend_income_total": 0, "foreign_tax_total": 0, "foreign_income_total": 0,
        },
    },
    {
        "id": "TC-CORP-08",
        "title": "인구감소지역_추가공제",
        "desc": "인구감소지역 소재 중소기업이 추가 공제율을 적용받는 시나리오",
        "basic": {
            "applicant_type": "C", "applicant_name": "주식회사 영주세라믹",
            "biz_reg_no": "890-12-34567", "tax_type": "CORP", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "23999",
            "hq_location": "경상북도 영주시 풍기읍 산업로 12", "capital_zone": "NON_CAPITAL",
            "depopulation_area": True, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 3000000000, "taxable_income": 500000000,
            "computed_tax": 75000000, "paid_tax": 75000000,
            "founding_date": "2008-09-15", "venture_yn": False, "rd_dept_yn": False,
            "claim_reason": "인구감소지역 고용증대 세액공제 추가공제 경정청구", "sincerity_target": False,
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 35.0, "youth_count": 8, "disabled_count": 1,
             "aged_count": 4, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 22, "excluded_count": 2, "total_salary": 1800000000, "social_insurance_paid": 180000000},
            {"year_type": "PREVIOUS", "total_regular": 30.0, "youth_count": 5, "disabled_count": 1,
             "aged_count": 4, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 20, "excluded_count": 2, "total_salary": 1500000000, "social_insurance_paid": 150000000},
        ],
        "deductions": [
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2024", "item_seq": 1,
             "base_amount": 0, "zone_type": "DEPOPULATION", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": None, "non_taxable_income": 20000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 30000000,
            "withholding_tax": 3000000, "determined_tax": 75000000,
            "dividend_income_total": 0, "foreign_tax_total": 0, "foreign_income_total": 0,
        },
    },
    {
        "id": "TC-CORP-09",
        "title": "이월공제_적용",
        "desc": "전기 미사용 세액공제를 이월하여 당기에 적용하는 시나리오",
        "basic": {
            "applicant_type": "C", "applicant_name": "주식회사 퓨처로보틱스",
            "biz_reg_no": "901-23-45678", "tax_type": "CORP", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "28111",
            "hq_location": "대구광역시 달서구 성서공단로 88", "capital_zone": "NON_CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 6000000000, "taxable_income": 900000000,
            "computed_tax": 171000000, "paid_tax": 171000000,
            "founding_date": "2016-04-01", "venture_yn": True, "rd_dept_yn": True,
            "claim_reason": "R&D 세액공제 이월분 + 당기분 경정청구", "sincerity_target": False,
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 55.0, "youth_count": 18, "disabled_count": 1,
             "aged_count": 3, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 33, "excluded_count": 4, "total_salary": 4000000000, "social_insurance_paid": 400000000},
            {"year_type": "PREVIOUS", "total_regular": 50.0, "youth_count": 15, "disabled_count": 1,
             "aged_count": 3, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 31, "excluded_count": 3, "total_salary": 3500000000, "social_insurance_paid": 350000000},
        ],
        "deductions": [
            {"item_category": "RD", "provision": "조특법10", "tax_year": "2024", "item_seq": 1,
             "base_amount": 1500000000, "zone_type": None, "asset_type": None,
             "rd_type": "GENERAL", "method": "CURRENT",
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
            {"item_category": "RD", "provision": "조특법10", "tax_year": "2023", "item_seq": 1,
             "base_amount": 1200000000, "zone_type": None, "asset_type": None,
             "rd_type": "GENERAL", "method": "CURRENT",
             "existing_applied": True, "existing_amount": 50000000, "carryforward_balance": 80000000},
        ],
        "financial": {
            "biz_income": None, "non_taxable_income": 40000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 80000000,
            "withholding_tax": 6000000, "determined_tax": 171000000,
            "dividend_income_total": 0, "foreign_tax_total": 0, "foreign_income_total": 0,
        },
    },
    {
        "id": "TC-CORP-10",
        "title": "중복배제_검증_시나리오",
        "desc": "창업감면과 고용증대 세액공제의 중복 적용 배제 규칙이 작동하는 시나리오",
        "basic": {
            "applicant_type": "C", "applicant_name": "주식회사 뉴스타트업",
            "biz_reg_no": "012-34-56789", "tax_type": "CORP", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "58210",
            "hq_location": "부산광역시 해운대구 센텀중앙로 97", "capital_zone": "NON_CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 4000000000, "taxable_income": 600000000,
            "computed_tax": 96000000, "paid_tax": 96000000,
            "founding_date": "2021-01-15", "venture_yn": True, "rd_dept_yn": True,
            "claim_reason": "창업감면 + 고용증대 세액공제 최적 조합 경정청구", "sincerity_target": False,
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 30.0, "youth_count": 15, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 15, "excluded_count": 2, "total_salary": 2000000000, "social_insurance_paid": 200000000},
            {"year_type": "PREVIOUS", "total_regular": 20.0, "youth_count": 10, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 10, "excluded_count": 1, "total_salary": 1200000000, "social_insurance_paid": 120000000},
        ],
        "deductions": [
            {"item_category": "STARTUP", "provision": "조특법6", "tax_year": "2024", "item_seq": 1,
             "base_amount": 600000000, "zone_type": "NON_CAPITAL", "asset_type": None,
             "rd_type": None, "method": None,
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2024", "item_seq": 1,
             "base_amount": 0, "zone_type": "NON_CAPITAL", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": None, "non_taxable_income": 30000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 40000000,
            "withholding_tax": 4000000, "determined_tax": 96000000,
            "dividend_income_total": 0, "foreign_tax_total": 0, "foreign_income_total": 0,
        },
    },
]

# ============================================================
# 개인 테스트 케이스 10건
# ============================================================
INC_CASES = [
    {
        "id": "TC-INC-01",
        "title": "개인사업자_고용증대_기본",
        "desc": "복식부기 의무 개인사업자가 고용증대 세액공제를 신청하는 기본 시나리오",
        "basic": {
            "applicant_type": "P", "applicant_name": "김세무",
            "biz_reg_no": "111-22-33333", "tax_type": "INC", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "56111",
            "hq_location": "경기도 수원시 팔달구 인계로 77", "capital_zone": "CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 1500000000, "taxable_income": 200000000,
            "computed_tax": 60920000, "paid_tax": 60920000,
            "founding_date": "2015-03-01", "venture_yn": False, "rd_dept_yn": False,
            "claim_reason": "고용증대 세액공제 경정청구", "sincerity_target": False,
            "bookkeeping_type": "DOUBLE_ENTRY",
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 10.0, "youth_count": 4, "disabled_count": 0,
             "aged_count": 1, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 5, "excluded_count": 1, "total_salary": 500000000, "social_insurance_paid": 50000000},
            {"year_type": "PREVIOUS", "total_regular": 7.0, "youth_count": 2, "disabled_count": 0,
             "aged_count": 1, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 4, "excluded_count": 0, "total_salary": 350000000, "social_insurance_paid": 35000000},
        ],
        "deductions": [
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2024", "item_seq": 1,
             "base_amount": 0, "zone_type": "CAPITAL", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": 200000000, "non_taxable_income": 5000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 25000000,
            "withholding_tax": 3000000, "determined_tax": 60920000,
            "inc_comprehensive_income": 200000000, "inc_deduction_total": 15000000,
        },
    },
    {
        "id": "TC-INC-02",
        "title": "개인사업자_RD_세액공제",
        "desc": "IT업종 개인사업자가 R&D 세액공제를 신청하는 시나리오",
        "basic": {
            "applicant_type": "P", "applicant_name": "이연구",
            "biz_reg_no": "222-33-44444", "tax_type": "INC", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "62010",
            "hq_location": "서울특별시 강남구 역삼로 150", "capital_zone": "CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 2000000000, "taxable_income": 300000000,
            "computed_tax": 97940000, "paid_tax": 97940000,
            "founding_date": "2013-08-01", "venture_yn": True, "rd_dept_yn": True,
            "claim_reason": "R&D 세액공제 경정청구", "sincerity_target": False,
            "bookkeeping_type": "DOUBLE_ENTRY",
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 15.0, "youth_count": 8, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 7, "excluded_count": 1, "total_salary": 900000000, "social_insurance_paid": 90000000},
            {"year_type": "PREVIOUS", "total_regular": 13.0, "youth_count": 7, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 6, "excluded_count": 1, "total_salary": 750000000, "social_insurance_paid": 75000000},
        ],
        "deductions": [
            {"item_category": "RD", "provision": "조특법10", "tax_year": "2024", "item_seq": 1,
             "base_amount": 500000000, "zone_type": None, "asset_type": None,
             "rd_type": "GENERAL", "method": "CURRENT",
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": 300000000, "non_taxable_income": 8000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 40000000,
            "withholding_tax": 5000000, "determined_tax": 97940000,
            "inc_comprehensive_income": 300000000, "inc_deduction_total": 18000000,
        },
    },
    {
        "id": "TC-INC-03",
        "title": "개인_창업감면_수도권외",
        "desc": "수도권 외 지역 창업 개인사업자가 세액감면을 신청하는 시나리오",
        "basic": {
            "applicant_type": "P", "applicant_name": "박창업",
            "biz_reg_no": "333-44-55555", "tax_type": "INC", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "63111",
            "hq_location": "광주광역시 북구 첨단과기로 123", "capital_zone": "NON_CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 800000000, "taxable_income": 100000000,
            "computed_tax": 24390000, "paid_tax": 24390000,
            "founding_date": "2022-01-10", "venture_yn": True, "rd_dept_yn": False,
            "claim_reason": "창업중소기업 세액감면 경정청구", "sincerity_target": False,
            "bookkeeping_type": "DOUBLE_ENTRY",
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 5.0, "youth_count": 3, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 2, "excluded_count": 0, "total_salary": 250000000, "social_insurance_paid": 25000000},
            {"year_type": "PREVIOUS", "total_regular": 3.0, "youth_count": 2, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 1, "excluded_count": 0, "total_salary": 150000000, "social_insurance_paid": 15000000},
        ],
        "deductions": [
            {"item_category": "STARTUP", "provision": "조특법6", "tax_year": "2024", "item_seq": 1,
             "base_amount": 100000000, "zone_type": "NON_CAPITAL", "asset_type": None,
             "rd_type": None, "method": None,
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": 100000000, "non_taxable_income": 3000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 10000000,
            "withholding_tax": 1000000, "determined_tax": 24390000,
            "inc_comprehensive_income": 100000000, "inc_deduction_total": 8000000,
        },
    },
    {
        "id": "TC-INC-04",
        "title": "성실신고_대상_개인사업자",
        "desc": "성실신고 확인 대상 고매출 개인사업자의 세액공제 시나리오",
        "basic": {
            "applicant_type": "P", "applicant_name": "최매출",
            "biz_reg_no": "444-55-66666", "tax_type": "INC", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "47190",
            "hq_location": "서울특별시 중구 명동길 14", "capital_zone": "CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 8000000000, "taxable_income": 500000000,
            "computed_tax": 171940000, "paid_tax": 171940000,
            "founding_date": "2005-06-01", "venture_yn": False, "rd_dept_yn": False,
            "claim_reason": "고용증대 세액공제 경정청구 (성실신고대상)", "sincerity_target": True,
            "bookkeeping_type": "DOUBLE_ENTRY",
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 25.0, "youth_count": 5, "disabled_count": 1,
             "aged_count": 3, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 16, "excluded_count": 2, "total_salary": 1500000000, "social_insurance_paid": 150000000},
            {"year_type": "PREVIOUS", "total_regular": 22.0, "youth_count": 4, "disabled_count": 1,
             "aged_count": 3, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 14, "excluded_count": 1, "total_salary": 1300000000, "social_insurance_paid": 130000000},
        ],
        "deductions": [
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2024", "item_seq": 1,
             "base_amount": 0, "zone_type": "CAPITAL", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": 500000000, "non_taxable_income": 10000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 80000000,
            "withholding_tax": 10000000, "determined_tax": 171940000,
            "inc_comprehensive_income": 500000000, "inc_deduction_total": 25000000,
        },
    },
    {
        "id": "TC-INC-05",
        "title": "개인_복합세액공제",
        "desc": "개인사업자가 고용증대 + R&D 세액공제를 복합 신청하는 시나리오",
        "basic": {
            "applicant_type": "P", "applicant_name": "정복합",
            "biz_reg_no": "555-66-77777", "tax_type": "INC", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "72192",
            "hq_location": "대전광역시 유성구 과학로 125", "capital_zone": "NON_CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 3000000000, "taxable_income": 400000000,
            "computed_tax": 134940000, "paid_tax": 134940000,
            "founding_date": "2017-02-01", "venture_yn": True, "rd_dept_yn": True,
            "claim_reason": "고용증대 + R&D 세액공제 복합 경정청구", "sincerity_target": False,
            "bookkeeping_type": "DOUBLE_ENTRY",
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 18.0, "youth_count": 8, "disabled_count": 0,
             "aged_count": 1, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 9, "excluded_count": 1, "total_salary": 1200000000, "social_insurance_paid": 120000000},
            {"year_type": "PREVIOUS", "total_regular": 14.0, "youth_count": 5, "disabled_count": 0,
             "aged_count": 1, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 8, "excluded_count": 1, "total_salary": 900000000, "social_insurance_paid": 90000000},
        ],
        "deductions": [
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2024", "item_seq": 1,
             "base_amount": 0, "zone_type": "NON_CAPITAL", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
            {"item_category": "RD", "provision": "조특법10", "tax_year": "2024", "item_seq": 1,
             "base_amount": 600000000, "zone_type": None, "asset_type": None,
             "rd_type": "GENERAL", "method": "CURRENT",
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": 400000000, "non_taxable_income": 12000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 55000000,
            "withholding_tax": 7000000, "determined_tax": 134940000,
            "inc_comprehensive_income": 400000000, "inc_deduction_total": 20000000,
        },
    },
    {
        "id": "TC-INC-06",
        "title": "간편장부_대상자",
        "desc": "간편장부 대상 소규모 개인사업자의 세액공제 신청 시나리오 (제한 있음)",
        "basic": {
            "applicant_type": "P", "applicant_name": "한소규",
            "biz_reg_no": "666-77-88888", "tax_type": "INC", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "56219",
            "hq_location": "인천광역시 남동구 논현로 55", "capital_zone": "CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 200000000, "taxable_income": 30000000,
            "computed_tax": 3564000, "paid_tax": 3564000,
            "founding_date": "2023-05-01", "venture_yn": False, "rd_dept_yn": False,
            "claim_reason": "고용증대 세액공제 경정청구", "sincerity_target": False,
            "bookkeeping_type": "SIMPLE",
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 3.0, "youth_count": 2, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 1, "excluded_count": 0, "total_salary": 120000000, "social_insurance_paid": 12000000},
            {"year_type": "PREVIOUS", "total_regular": 1.0, "youth_count": 1, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 0, "excluded_count": 0, "total_salary": 40000000, "social_insurance_paid": 4000000},
        ],
        "deductions": [
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2024", "item_seq": 1,
             "base_amount": 0, "zone_type": "CAPITAL", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": 30000000, "non_taxable_income": 1000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 0,
            "withholding_tax": 500000, "determined_tax": 3564000,
            "inc_comprehensive_income": 30000000, "inc_deduction_total": 5000000,
        },
    },
    {
        "id": "TC-INC-07",
        "title": "개인_벤처기업_확인",
        "desc": "벤처기업 확인 개인사업자가 R&D 세액공제를 신청하는 시나리오",
        "basic": {
            "applicant_type": "P", "applicant_name": "오벤처",
            "biz_reg_no": "777-88-99999", "tax_type": "INC", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "58219",
            "hq_location": "경기도 성남시 분당구 판교역로 235", "capital_zone": "CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 1000000000, "taxable_income": 150000000,
            "computed_tax": 42140000, "paid_tax": 42140000,
            "founding_date": "2020-09-01", "venture_yn": True, "rd_dept_yn": True,
            "claim_reason": "벤처기업 R&D 세액공제 경정청구", "sincerity_target": False,
            "bookkeeping_type": "DOUBLE_ENTRY",
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 8.0, "youth_count": 5, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 3, "excluded_count": 0, "total_salary": 600000000, "social_insurance_paid": 60000000},
            {"year_type": "PREVIOUS", "total_regular": 6.0, "youth_count": 4, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 2, "excluded_count": 0, "total_salary": 420000000, "social_insurance_paid": 42000000},
        ],
        "deductions": [
            {"item_category": "RD", "provision": "조특법10", "tax_year": "2024", "item_seq": 1,
             "base_amount": 300000000, "zone_type": None, "asset_type": None,
             "rd_type": "GENERAL", "method": "CURRENT",
             "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": 150000000, "non_taxable_income": 5000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 15000000,
            "withholding_tax": 2000000, "determined_tax": 42140000,
            "inc_comprehensive_income": 150000000, "inc_deduction_total": 12000000,
        },
    },
    {
        "id": "TC-INC-08",
        "title": "인구감소지역_개인사업자",
        "desc": "인구감소지역 소재 개인사업자의 고용증대 추가공제 시나리오",
        "basic": {
            "applicant_type": "P", "applicant_name": "강지방",
            "biz_reg_no": "888-99-00000", "tax_type": "INC", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "10611",
            "hq_location": "강원특별자치도 삼척시 근덕면 산업로 10", "capital_zone": "NON_CAPITAL",
            "depopulation_area": True, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 600000000, "taxable_income": 80000000,
            "computed_tax": 17060000, "paid_tax": 17060000,
            "founding_date": "2010-04-01", "venture_yn": False, "rd_dept_yn": False,
            "claim_reason": "인구감소지역 고용증대 세액공제 추가공제 경정청구", "sincerity_target": False,
            "bookkeeping_type": "DOUBLE_ENTRY",
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 8.0, "youth_count": 2, "disabled_count": 1,
             "aged_count": 2, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 3, "excluded_count": 1, "total_salary": 350000000, "social_insurance_paid": 35000000},
            {"year_type": "PREVIOUS", "total_regular": 6.0, "youth_count": 1, "disabled_count": 1,
             "aged_count": 2, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 2, "excluded_count": 0, "total_salary": 260000000, "social_insurance_paid": 26000000},
        ],
        "deductions": [
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2024", "item_seq": 1,
             "base_amount": 0, "zone_type": "DEPOPULATION", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": 80000000, "non_taxable_income": 2000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 5000000,
            "withholding_tax": 1000000, "determined_tax": 17060000,
            "inc_comprehensive_income": 80000000, "inc_deduction_total": 7000000,
        },
    },
    {
        "id": "TC-INC-09",
        "title": "경정청구_기한초과",
        "desc": "경정청구 기한(5년)이 지난 귀속연도에 대해 신청하여 기한 검증이 실패하는 시나리오",
        "basic": {
            "applicant_type": "P", "applicant_name": "윤기한",
            "biz_reg_no": "999-00-11111", "tax_type": "INC", "tax_year": "2018",
            "corp_size": "SMALL", "industry_code": "46410",
            "hq_location": "서울특별시 종로구 세종대로 209", "capital_zone": "CAPITAL",
            "depopulation_area": False, "fiscal_start": "2018-01-01", "fiscal_end": "2018-12-31",
            "revenue": 1200000000, "taxable_income": 180000000,
            "computed_tax": 47540000, "paid_tax": 47540000,
            "founding_date": "2008-01-15", "venture_yn": False, "rd_dept_yn": False,
            "claim_reason": "고용증대 세액공제 경정청구 (2018년 귀속)", "sincerity_target": False,
            "bookkeeping_type": "DOUBLE_ENTRY",
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 12.0, "youth_count": 4, "disabled_count": 0,
             "aged_count": 1, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 7, "excluded_count": 1, "total_salary": 700000000, "social_insurance_paid": 70000000},
            {"year_type": "PREVIOUS", "total_regular": 10.0, "youth_count": 3, "disabled_count": 0,
             "aged_count": 1, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 6, "excluded_count": 0, "total_salary": 580000000, "social_insurance_paid": 58000000},
        ],
        "deductions": [
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2018", "item_seq": 1,
             "base_amount": 0, "zone_type": "CAPITAL", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": 180000000, "non_taxable_income": 5000000,
            "loss_carryforward_total": 0, "interim_prepaid_tax": 20000000,
            "withholding_tax": 3000000, "determined_tax": 47540000,
            "inc_comprehensive_income": 180000000, "inc_deduction_total": 12000000,
        },
    },
    {
        "id": "TC-INC-10",
        "title": "개인_결손금_이월공제",
        "desc": "전기 결손금이 있어 과세표준이 줄어든 상태에서 세액공제를 신청하는 시나리오",
        "basic": {
            "applicant_type": "P", "applicant_name": "송결손",
            "biz_reg_no": "000-11-22222", "tax_type": "INC", "tax_year": "2024",
            "corp_size": "SMALL", "industry_code": "73111",
            "hq_location": "서울특별시 마포구 월드컵북로 396", "capital_zone": "CAPITAL",
            "depopulation_area": False, "fiscal_start": "2024-01-01", "fiscal_end": "2024-12-31",
            "revenue": 900000000, "taxable_income": 50000000,
            "computed_tax": 5940000, "paid_tax": 5940000,
            "founding_date": "2019-07-01", "venture_yn": False, "rd_dept_yn": False,
            "claim_reason": "고용증대 세액공제 경정청구 (결손금 이월 상태)", "sincerity_target": False,
            "bookkeeping_type": "DOUBLE_ENTRY",
        },
        "employees": [
            {"year_type": "CURRENT", "total_regular": 7.0, "youth_count": 3, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 4, "excluded_count": 0, "total_salary": 420000000, "social_insurance_paid": 42000000},
            {"year_type": "PREVIOUS", "total_regular": 5.0, "youth_count": 2, "disabled_count": 0,
             "aged_count": 0, "career_break_count": 0, "north_defector_count": 0,
             "general_count": 3, "excluded_count": 0, "total_salary": 300000000, "social_insurance_paid": 30000000},
        ],
        "deductions": [
            {"item_category": "EMPLOYMENT", "provision": "조특법30의4", "tax_year": "2024", "item_seq": 1,
             "base_amount": 0, "zone_type": "CAPITAL", "asset_type": None, "rd_type": None,
             "method": None, "existing_applied": False, "existing_amount": 0, "carryforward_balance": 0},
        ],
        "financial": {
            "biz_income": 120000000, "non_taxable_income": 3000000,
            "loss_carryforward_total": 70000000,
            "loss_carryforward_detail": '[{"year":"2022","amount":30000000},{"year":"2023","amount":40000000}]',
            "interim_prepaid_tax": 0,
            "withholding_tax": 1000000, "determined_tax": 5940000,
            "inc_comprehensive_income": 50000000, "inc_deduction_total": 8000000,
            "current_year_loss": 0,
        },
    },
]


def build_request_json(case, request_date="2025-02-10"):
    """엔티티 구조에 맞는 JSON 데이터를 생성한다."""
    basic = case["basic"]
    req_id = case["id"].replace("-", "")
    data = {
        "request": {
            "req_id": req_id,
            "applicant_type": basic["applicant_type"],
            "applicant_id": basic["biz_reg_no"],
            "applicant_name": basic["applicant_name"],
            "tax_type": basic["tax_type"],
            "tax_year": basic["tax_year"],
            "request_date": request_date,
            "seq_no": 1,
            "request_status": "RECEIVED",
            "request_source": "TEST",
        },
        "basic": {
            "req_id": req_id,
            "request_date": request_date,
            "tax_type": basic["tax_type"],
            "applicant_name": basic["applicant_name"],
            "biz_reg_no": basic["biz_reg_no"],
            "corp_size": basic["corp_size"],
            "industry_code": basic["industry_code"],
            "hq_location": basic["hq_location"],
            "capital_zone": basic["capital_zone"],
            "depopulation_area": basic["depopulation_area"],
            "tax_year": basic["tax_year"],
            "fiscal_start": basic["fiscal_start"],
            "fiscal_end": basic["fiscal_end"],
            "revenue": basic["revenue"],
            "taxable_income": basic["taxable_income"],
            "computed_tax": basic["computed_tax"],
            "paid_tax": basic["paid_tax"],
            "founding_date": basic["founding_date"],
            "venture_yn": basic["venture_yn"],
            "rd_dept_yn": basic["rd_dept_yn"],
            "claim_reason": basic["claim_reason"],
            "sincerity_target": basic.get("sincerity_target", False),
            "bookkeeping_type": basic.get("bookkeeping_type"),
        },
        "employees": [
            {**emp, "req_id": req_id} for emp in case["employees"]
        ],
        "deductions": [
            {**ded, "req_id": req_id} for ded in case["deductions"]
        ],
        "financial": {
            "req_id": req_id,
            **case["financial"],
        },
    }
    return data


def write_excel(case, filepath):
    """테스트 케이스를 엑셀 파일로 작성한다."""
    wb = Workbook()

    def style_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    def write_row(ws, row, values):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = THIN_BORDER

    # --- 시트1: 개요 ---
    ws_overview = wb.active
    ws_overview.title = "개요"
    ws_overview.column_dimensions["A"].width = 20
    ws_overview.column_dimensions["B"].width = 60
    overview_data = [
        ("테스트케이스 ID", case["id"]),
        ("테스트케이스명", case["title"]),
        ("설명", case["desc"]),
        ("세금유형", case["basic"]["tax_type"]),
        ("신청자유형", "법인" if case["basic"]["applicant_type"] == "C" else "개인"),
        ("귀속연도", case["basic"]["tax_year"]),
    ]
    style_header(ws_overview, ["항목", "값"])
    for i, (k, v) in enumerate(overview_data, 2):
        write_row(ws_overview, i, [k, v])

    # --- 시트2: 기본정보 ---
    ws_basic = wb.create_sheet("기본정보")
    basic = case["basic"]
    basic_headers = ["항목", "값"]
    style_header(ws_basic, basic_headers)
    basic_rows = [
        ("신청자명", basic["applicant_name"]),
        ("사업자등록번호", basic["biz_reg_no"]),
        ("세금유형", basic["tax_type"]),
        ("귀속연도", basic["tax_year"]),
        ("기업규모", basic["corp_size"]),
        ("업종코드", basic["industry_code"]),
        ("본점소재지", basic["hq_location"]),
        ("수도권구분", basic["capital_zone"]),
        ("인구감소지역", "Y" if basic["depopulation_area"] else "N"),
        ("사업연도시작", basic["fiscal_start"]),
        ("사업연도종료", basic["fiscal_end"]),
        ("매출액", basic["revenue"]),
        ("과세표준", basic["taxable_income"]),
        ("산출세액", basic["computed_tax"]),
        ("기납부세액", basic["paid_tax"]),
        ("설립일", basic["founding_date"]),
        ("벤처확인", "Y" if basic["venture_yn"] else "N"),
        ("R&D부서", "Y" if basic["rd_dept_yn"] else "N"),
        ("청구사유", basic["claim_reason"]),
    ]
    if basic.get("bookkeeping_type"):
        basic_rows.append(("장부유형", basic["bookkeeping_type"]))
    ws_basic.column_dimensions["A"].width = 20
    ws_basic.column_dimensions["B"].width = 50
    for i, (k, v) in enumerate(basic_rows, 2):
        write_row(ws_basic, i, [k, v])

    # --- 시트3: 고용정보 ---
    ws_emp = wb.create_sheet("고용정보")
    emp_headers = [
        "연도구분", "상시근로자수", "청년수", "장애인수", "고령자수",
        "경력단절여성수", "북한이탈주민수", "일반근로자수", "제외근로자수",
        "총급여", "사회보험납부액"
    ]
    style_header(ws_emp, emp_headers)
    for i, emp in enumerate(case["employees"], 2):
        write_row(ws_emp, i, [
            emp["year_type"], emp["total_regular"], emp["youth_count"],
            emp["disabled_count"], emp["aged_count"], emp["career_break_count"],
            emp["north_defector_count"], emp["general_count"], emp["excluded_count"],
            emp["total_salary"], emp["social_insurance_paid"],
        ])
    for col in range(1, len(emp_headers) + 1):
        ws_emp.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 16

    # --- 시트4: 공제항목 ---
    ws_ded = wb.create_sheet("공제항목")
    ded_headers = [
        "항목분류", "조항", "귀속연도", "순번", "기준금액",
        "지역유형", "자산유형", "R&D유형", "방법",
        "기존적용여부", "기존적용금액", "이월잔액"
    ]
    style_header(ws_ded, ded_headers)
    for i, ded in enumerate(case["deductions"], 2):
        write_row(ws_ded, i, [
            ded["item_category"], ded["provision"], ded["tax_year"],
            ded["item_seq"], ded["base_amount"],
            ded.get("zone_type", ""), ded.get("asset_type", ""),
            ded.get("rd_type", ""), ded.get("method", ""),
            "Y" if ded.get("existing_applied") else "N",
            ded.get("existing_amount", 0), ded.get("carryforward_balance", 0),
        ])
    for col in range(1, len(ded_headers) + 1):
        ws_ded.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 16

    # --- 시트5: 재무정보 ---
    ws_fin = wb.create_sheet("재무정보")
    fin = case["financial"]
    fin_headers = ["항목", "값"]
    style_header(ws_fin, fin_headers)
    fin_rows = []
    for k, v in fin.items():
        if v is not None:
            label = {
                "biz_income": "사업소득금액",
                "non_taxable_income": "비과세소득",
                "loss_carryforward_total": "이월결손금합계",
                "loss_carryforward_detail": "이월결손금상세",
                "interim_prepaid_tax": "중간예납세액",
                "withholding_tax": "원천징수세액",
                "determined_tax": "결정세액",
                "dividend_income_total": "배당소득합계",
                "dividend_exclusion_detail": "배당소득제외상세",
                "foreign_tax_total": "외국납부세액",
                "foreign_income_total": "해외소득합계",
                "inc_comprehensive_income": "종합소득금액",
                "inc_deduction_total": "소득공제합계",
                "current_year_loss": "당기결손금",
            }.get(k, k)
            fin_rows.append((label, v))
    ws_fin.column_dimensions["A"].width = 25
    ws_fin.column_dimensions["B"].width = 30
    for i, (k, v) in enumerate(fin_rows, 2):
        write_row(ws_fin, i, [k, v])

    wb.save(filepath)


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    corp_dir = os.path.join(base_dir, "corp")
    inc_dir = os.path.join(base_dir, "individual")
    os.makedirs(corp_dir, exist_ok=True)
    os.makedirs(inc_dir, exist_ok=True)

    # 법인 케이스
    for case in CORP_CASES:
        name = case["id"]
        json_data = build_request_json(case)
        json_path = os.path.join(corp_dir, f"{name}.json")
        xlsx_path = os.path.join(corp_dir, f"{name}.xlsx")

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        write_excel(case, xlsx_path)
        print(f"[CORP] {name} -> {json_path}, {xlsx_path}")

    # 개인 케이스
    for case in INC_CASES:
        name = case["id"]
        json_data = build_request_json(case)
        json_path = os.path.join(inc_dir, f"{name}.json")
        xlsx_path = os.path.join(inc_dir, f"{name}.xlsx")

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        write_excel(case, xlsx_path)
        print(f"[INC]  {name} -> {json_path}, {xlsx_path}")

    print(f"\n총 {len(CORP_CASES) + len(INC_CASES)}건 테스트 케이스 생성 완료!")


if __name__ == "__main__":
    main()
